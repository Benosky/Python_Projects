#!/usr/bin/env python2
# -*- coding: utf-8 -*-
"""
Created on Sun Dec 16 16:46:02 2018

@author: benjaminumeh
"""

    
import pandas as pd
#from pandas import ExcelWriter
#import itertools
import time
#import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
import scipy.optimize as spo


#def sku_per_locn
def locn_per_sku(md30):
    """ """
    """        
    import pandas as pd
    from pandas import ExcelWriter
    #import itertools
    import time
    #import matplotlib.pyplot as plt
    import numpy as np
    from openpyxl import load_workbook
    """
    
    #locn = "All Purpose Pharmacy"
    #Read the excel sheet and assign it to data 
    data = pd.read_excel(md30, sheet_name='Sheet1')
    #Fill NaN with zero to prevent groupby from excluding some columns
    md = data.fillna(0)
    #Delete rows where prev.opening and closing are both zeroes
    #md = md[(md[["prev. opening","closing"]] != 0).all(axis=1)]
    #Create a 'month_year' column
    md['month_year'] = md.date.dt.to_period('M')
    #Convert the 'Price' column to numeric to aid their multiplication
    #md["Price"] = pd.to_numeric(md["Price"],errors='coerce')
    #md["Average Inventory"] = (md["prev. opening"] + md["closing"])/2.0
    #md["Inventory Amount"] = md["Average Inventory"] * md["Price"]
    #md["Inventory Finance Cost"] = (md["Inventory Amount"]*((30/float(365))*0.32))#md["Inventory Amount"]
    #Set the "SKU" column as the index
    mdsku = md.set_index('SKU')
    #Load the data into a dataframe
    mdf = pd.DataFrame(mdsku)
    #Sort the dataframe according to the Location
    mdf = mdf.sort_values('SKU',inplace=False)
    #Extract the list of unique index values and convert it from list of unicodes to list of strings
    #ind_sk = map(str,list(mdf.index.unique()))
    ind_sk = list(mdf.index.unique())
    #Get the list of the unique values in the month_year column
    
    #report = report.set_index('month_year')
    #Create empty lists to hold the dataframes of each location metrics
    loc_per_SKU = [] 
    prod_names = []
    #avg_inv_amt = []
    #inv_fin_cost = []
    #loc_sku = []
    #cost_per_sku = []
    for l in ind_sk:
        
        #Filter the dataframe by the location
        dt = mdf.loc[[l]]
        #Get the product name
        prdnm = dt["product"].iloc[0]
        #Sort the dataframe by "location" and "date"
        dt = dt.sort_values(by=['date','location'])
        #Reset the index
        dt = dt.reset_index()
        #Set the year-month as the index
        dt = dt.set_index('month_year')
        #extract the unique year months into a list
        #yr_months = map(str,list(dt.index.unique()))#report.month_year.unique()
        yr_months = list(dt.index.unique())#report.month_year.unique()
        #Create empty DataFrame to hold location data
        loc_df = pd.DataFrame()
        for y in yr_months:
            dt_m = dt.loc[[y]]
            #Select the required columns from the dataframe
            dt_m = dt_m[["location",'date']]
            dt_m.columns = ['loc_per('+str(l)+')','date']
            #group the dataframe by "locationsku" and compute the mean Inventory Amount and Inventory Finance Cost
            #dt = dt.groupby('SKU',as_index=False)[["Inventory Amount","Inventory Finance Cost"]].mean()
            dt_m = (dt_m.groupby('loc_per('+str(l)+')').agg({'loc_per('+str(l)+')':'last', 'date': 'last'}))#,'inv_cst('+l+')': 'mean','date':'last'}))
                    #.rename(columns={'SKU':'SKU Count'}))       
            dt_m['loc_per('+str(l)+')'] = 1
            dt_m = dt_m.reset_index(drop=True)
            #dt_m = dt_m.set_index('date')
            dt_m.index = dt_m.date.dt.to_period('M') #dt_m['date']
            
            dt_fm = dt_m.drop(['date'], axis = 1)
            
            #dt_ym = dt_ym.sum()
            dt_ym = dt_fm.sum()
            dt_ym = pd.DataFrame(dt_ym)
            dt_ym = dt_ym.T 
            dt_ym.index = [dt_fm.index[0]]
            #dt_ym['Cst/SKU('+l+')'] = dt_ym['inv_cst('+l+')']/dt_ym['sku_cnt('+l+')']
            loc_df = loc_df.append(dt_ym)
            
        #loc_sku.append(loc_df)
        loc_per_SKU.append(loc_df)#[['location']])  
        prod_names.append(prdnm)
        #avg_inv_amt.append(loc_df[['Inv_amt('+l+')']])
        #inv_fin_cost.append(loc_df[['inv_cst('+l+')']])
        #cost_per_sku.append(loc_df[['Cst/SKU('+l+')']])
    #hist_loc_sku = pd.concat(loc_sku, axis=1, sort=False)
    hist_loc_psku = pd.concat(loc_per_SKU, axis=1, sort=False)
    
    #Trim the column names
    hlocpsku = list(hist_loc_psku.columns.values)
    lccnt_sku = [x[8:-1] for x in hlocpsku]
    hist_loc_psku.columns = lccnt_sku
    
    hist_loc_psku = hist_loc_psku.T
    hist_loc_psku.index.name = "SKU"
    hist_loc_psku = hist_loc_psku.reset_index(inplace=False)
    hist_loc_psku.index.name = "Index"
    hist_loc_psku.index += 1
    hist_loc_psku = hist_loc_psku.round(0)
    hist_loc_psku.insert(1, "Product Name", prod_names, allow_duplicates=False)

    book = load_workbook('KPI_locationsku.xlsx')
    writer = pd.ExcelWriter('KPI_locationsku.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    #data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])
    hist_loc_psku.to_excel(writer,sheet_name='Hist_Locn_Per_SKU',index=True)
    #hist_cost_per_sku.to_excel(writer,sheet_name='Hist_Invt_Cost_Per_SKU',index=True)
    #hist_sku_ploc.to_excel(writer,sheet_name='Hist_SKU_Per_Location',index=True)
    #hist_invamt_ploc.to_excel(writer,sheet_name='Hist_Inventory_Amount_Locn',index=True)
    #hist_invcst_ploc.to_excel(writer,sheet_name='Hist_Inventory_Fin_Cost_Per_Loc',index=True)
    writer.save()
    locsku_result = writer.save()#_kpi.save()
    return locsku_result

locn_per_sku('md_30.xlsx')
    

# Import the required modules

import numpy as np
import pandas as pd
from pandas_datareader import data as pdr
import fix_yahoo_finance as yf
#from Compute_KPI_OutputClass import Output
import scipy.optimize as spo
import random
import time
import datetime 
# Import matplotlib
import matplotlib.pyplot as plt 

yf.pdr_override()
from sklearn.ensemble import RandomForestClassifier
from sklearn.ensemble import GradientBoostingRegressor
from sklearn import tree
from sklearn import preprocessing
from sklearn.datasets import make_classification
from sklearn.metrics import mean_squared_error
from sklearn.datasets import make_friedman1
from sklearn.ensemble import AdaBoostClassifier
from sklearn.ensemble import RandomForestRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.ensemble import AdaBoostRegressor
#from Compute_KPI_OutputClass import Output
from datetime import timedelta
import calendar
from sklearn.model_selection import cross_val_score
from sklearn.metrics import mean_squared_error
from datetime import datetime

def demand_fxn(md30,kpi,price_calc):
    # = pd.DataFrame()
    #Import historical monthly qty sold data
    mnly_qty_sld_df = pd.read_excel(kpi, sheet_name='Historical_Monthly_Qty_Sold_SKU')
    #get the list of dates
    dates = list(mnly_qty_sld_df.columns.values)[3:]
    #Sort the dataframe by SKU
    #mnly_qty_sld_df = mnly_qty_sld_df.sort_values(by=['SKU'], inplace = True)
    mnly_qty_sld_df.sort_values(by=['SKU'], inplace = True)
    #get the series of the Products
    #products = map(str, list(mnly_qty_sld_df["Products"]))
    
    #get the series of the SKUs  
    sku = list(mnly_qty_sld_df["SKU"])
    
    #Reset the index and drop the former index
    ms_df = mnly_qty_sld_df.reset_index(drop=True)
    #Set the SKU column as the index
    ms_df = ms_df.set_index('SKU')
    
    #Import Historical Locations per SKU
    locn_per_sku_df = pd.read_excel(kpi, sheet_name='Hist_Locn_Per_SKU')
    #month_year = map(str,list(locn_per_sku_df.columns.values))[3:]
    #Get the required range of dates
    month_year = locn_per_sku_df.columns.values[3:]
    #month_year = [datetime.strptime(x, '%Y-%m') for x in month_year]
    #month_year = month_year.date.dt.to_period('M')
    
    #Reset the index and replace it with the SKUs
    ls_df = locn_per_sku_df.reset_index(drop=True)
    ls_df = ls_df.set_index('SKU')
    
    #Import Master Data 3.0
    sell_pr_df = pd.read_excel(md30, sheet_name='Sheet1')
    #Fill NaN with zero to prevent groupby from excluding some columns
    spd = sell_pr_df.fillna(0)
    #Create a month_year column
    spd['month_year'] = spd.date.dt.to_period('M')
    #Concatenate the month_year and the SKU number to get the sku_month_year column
    spd['sku_month_year'] = spd["SKU"].astype(str) +'-'+ spd['month_year'].astype(str)#+'-'+md['product']
    #Set the SKU column as the index
    spd = spd.set_index('SKU')
    #Import the buy price sheet
    bpr = pd.read_excel(price_calc, sheet_name='Sheet1')
    #Reset the index and replace it with the SKUs
    bpr = bpr.reset_index(drop=True)
    bpr = bpr.set_index('ID')
    bpr = bpr.fillna(method='ffill', axis=1)
    bpr.iloc[:,-1] = pd.to_numeric(bpr.iloc[:,-1],errors='coerce')
    
    demand_forecast = []
    var_coef = []
    mod_intercept = []
    price_range = []
    current_features = []
    feat_indx = []
    products = []
    for i in sku:
        if i in list(bpr.index):
            #get the Product name
            product = str(bpr.loc[i,"Product"])
            #Filter the data of the selected SKU for just the date ranges
            mnly_qty_sld_per_sku = pd.DataFrame(ms_df.loc[i,  dates[0]:])
            #name the index as 'date'
            mnly_qty_sld_per_sku.index.name = 'date'
            #Reset the index
            mnly_qty_sld_per_sku = mnly_qty_sld_per_sku.reset_index()
            #Convert the date column to datetime type
            mnly_qty_sld_per_sku['date'] = pd.to_datetime(mnly_qty_sld_per_sku['date'], errors='coerce')
            #Set the index to year_month period
            mnly_qty_sld_per_sku.index = mnly_qty_sld_per_sku["date"].dt.to_period('M')
            #rename the index as "month_year"
            mnly_qty_sld_per_sku.index.name = "month_year"
            #Drop the date column
            mnly_qty_sld_per_sku = mnly_qty_sld_per_sku.drop("date", axis=1)
            #Label the single column as "Monthly_Sales"
            mnly_qty_sld_per_sku["Monthly_Sales"] = pd.to_numeric(mnly_qty_sld_per_sku.iloc[:,0])
            mnly_qty_sld_per_sku = mnly_qty_sld_per_sku.drop(i, axis=1)
            #Fill NaN with 0
            mnly_qty_sld_per_sku = mnly_qty_sld_per_sku.fillna(0)
            #load into DataFrame
            mnly_qty_sld_per_sku = pd.DataFrame(mnly_qty_sld_per_sku)
            #main_df = main_df.append(mnly_sales_per_sku)
            #main_df
            #Filter the data of the selected SKU for just the date ranges
            lc_per_sku = pd.DataFrame(ls_df.loc[i,  month_year[0]:])
            
            #Create loc_per_SKU_shift column 
            #lc_per_sku["loc_per_SKU_shift"] = lc_per_sku[sku[i]].shift(1)
            #Drop the SKU number column and fill NaN with 0
            #lc_per_sku = lc_per_sku.drop(sku[1], axis=1)
            lc_per_sku.columns = ["Locations_per_sku"]
            lc_per_sku["loc_per_SKU_shift"] = lc_per_sku["Locations_per_sku"].shift(1)
            lc_per_sku = lc_per_sku.fillna(0)
            
            #Load to DataFrame
            lc_per_sku = pd.DataFrame(lc_per_sku)
            #Convert the month_year from type object to type DateTime
            lc_per_sku['date'] = [datetime.strptime(x, '%Y-%m') for x in month_year]
            #Reset index
            lc_per_sku = lc_per_sku.reset_index(drop=True)
            #Convert the date to month_year of type Period to make it concatenable with other dataframes
            lc_per_sku.index = lc_per_sku["date"].dt.to_period('M')
            #lc_per_sku = lc_per_sku.set_index('date')
            #Drop the date column and name the index as month_year
            lc_per_sku = lc_per_sku.drop("date", axis=1)
            lc_per_sku.index.name = 'month_year'
        
        
            #Filter the data for the selected SKU and drop NaN rows
            spd2 = spd.loc[[i]]
            sp_df = spd2.dropna()
            #Groupby and set month_year as the index
            sp_df = (sp_df.groupby('sku_month_year').agg({'month_year':'last', 'Price': 'last'}))#,'inv_cst('+l+')': 'mean','date':'last'}))
            sp_df = sp_df.set_index('month_year')
            #Create sell_price_shift column 
            sp_df["Price"] = sp_df["Price"].replace({0:np.nan})
            sp_df["Price"] = sp_df["Price"].fillna(method='ffill')
            sp_df["Price_shift"] = sp_df['Price'].shift(1)
            #sp_df = sp_df.drop("Price", axis=1)
            sp_df = sp_df.fillna(0)
            sp_df = pd.DataFrame(sp_df)
            
            df_list = [sp_df,mnly_qty_sld_per_sku,lc_per_sku]
            dd_df = pd.concat(df_list, axis=1, sort = True)
            #dd_df["Seasons"] = [x[6:] for x in list(dd_df.index.values)]
            #dd_df["Seasons"] = list(dd_df.index.values)#.apply(lambda x : x[6:])
            dd_df = dd_df.reset_index()
            dd_df.index = dd_df["month_year"]#.to_datetime()
            #Convert index from Period to datetime
            dd_df.index = dd_df.index.to_timestamp(freq=None, how='end') 
            
            #dd_df["Seasons"] = dd_df["month_year"].apply(lambda x : x[6:])
            #Convert column 'month_year' to string and slice out the year part
            dd_df["Seasons"] = [x[5:] for x in list(dd_df["month_year"].astype(str))]
            dd_df["Seasons"] = pd.to_numeric(dd_df["Seasons"])
            dd_df["Seasons_shift"] = dd_df["Seasons"].shift(1)
            dd_df["Seasons_shift"] = pd.to_numeric(dd_df["Seasons_shift"])
            
            
            trend = range(1,len(dd_df)+1)
            dd_df["Trend"] = pd.to_numeric(trend)# trend
            dd_df["Trend_shift"] = dd_df["Trend"].shift(1)
            dd_df = dd_df.drop(["month_year"], axis=1)
            #dd_df["Price"] = dd_df["Price"].replace({0:np.nan})
            #dd_df["Price"] = pd.to_numeric(dd_df["Price"].fillna(method='ffill'))
            #dd_df["Price"] = (dd_df["Price"].fillna(method='ffill'))
            #Get the price ceiling multiplier, mark-up and mark-down
            #ceil_multiplier = bpr.loc[i][-3:].std()
            m_up = 0.03
            m_dn = 0.0075
            pr = (bpr.loc[i][-1] - bpr.loc[i][-1]*m_dn, bpr.loc[i][-1] + (bpr.loc[i][-1]*m_up))# + (bpr.loc[i][-1]*ceil_multiplier)) # dd_df["Price"].iloc[-1] + (dd_df["Price"].iloc[-1] * ceil_multiplier)
            #drop the unrequired columns
            dd_df = dd_df.drop(["Price","Seasons","Trend","Locations_per_sku"], axis=1)
            
            #dd_df = pd.date_range(start='1/1/2018', periods=5, freq='M')
            
            # Extract exogenous and endogenous variables
            endogenous = dd_df['Monthly_Sales'][1:-1]
            
            features = dd_df.drop(['Monthly_Sales'], axis = 1)#[-1:]
            features = features.fillna(0)
            current_feat = features[-1:]
            exogenous = features[1:-1]
            
            #List of exogenous variables for later use
            exogenous_list = list(exogenous.columns)
            #Out[714]: ['Price', 'Locations_per_sku', 'Seasons', 'Trend']
            #print exogenous_list
            #Check arrays shapes
            endogenous.shape
            exogenous.shape
            current_feat.shape
            
            """
            # Convert to numpy arrays
            exogenous = np.array(exogenous)
            endogenous = np.array(endogenous)
            current_feat = np.array(current_feat)#
            current_feat
            """
            
            #define a size variable
            size = int(len(exogenous) * 0.75)
            
            #divide the data into training and testing components
            #train_exogenous, test_exogenous = exogenous[0:size], exogenous[size:len(exogenous)]
            #train_endogenous, test_endogenous = endogenous[0:size], endogenous[size:len(endogenous)]
            train_exogenous, test_exogenous = exogenous, exogenous#[0:size], exogenous[size:len(exogenous)]
            train_endogenous, test_endogenous = endogenous, endogenous#[0:size], endogenous[size:len(endogenous)]
            
            #def ml_reg((n_est1,n_est2)):
                    
            #Set up the Random Forest Classifier
            #rfr = RandomForestRegressor(n_estimators=240, max_depth=2,random_state=0)
            
            #Set up the Adaboost Classifier using the above RandomForestClassifier ensemble model as the base_estimator
            abr = AdaBoostRegressor(base_estimator=None, n_estimators=300, learning_rate=1.0,  random_state=None)
            
            #scores = cross_val_score(abr, exogenous, endogenous, cv=5)
            #scores.mean()  
            #[, sample_weight])	
            
            #est = abc.fit(train_exogenous,train_endogenous)
            abr.fit(train_exogenous,train_endogenous)
            #return -abr.score(exogenous, endogenous)
            abr.score(exogenous, endogenous)#Returns the coefficient of determination R^2 of the prediction. 
                                            #How well the model predicts the future outcome. The model will predict accurately 99% of the time
            
            # Print list of exogenous variables 
            #print abr.feature_importances_
            
            ml_res = int(abr.predict(current_feat))
            
            from sklearn.linear_model import LinearRegression
            #X = np.array([[1, 1], [1, 2], [2, 2], [2, 3]])
            # y = 1 * x_0 + 2 * x_1 + 3
            #y = np.dot(X, np.array([1, 2])) + 3
            reg = LinearRegression().fit(exogenous, endogenous)
            #reg.score(exogenous, endogenous)
            coeff = list(reg.coef_.round(2))
            intercept = reg.intercept_ 
            """
            from statsmodels.tsa.arima_model import ARIMA
        
            
            model = ARIMA(endogenous,exog = exogenous, order=(2,1,1),missing = 'drop')
            #start_params = [0, 0, 1.]
            model_fit = model.fit()#(disp=-1)
            #start_params = [0, 0, 1.]
            #results = mod.fit(start_params)
            coeff = list(results.params)
            #model_fit = model_fit
            #forecast_d = model_fit.predict(start=len(product), end=len(product), typ='levels',dynamic=False) 
            #next_wk_forecast = forecast_d.round(0)
            """
            
            """
            import statsmodels.api as sm
        
            mod = sm.tsa.statespace.SARIMAX(endogenous, exog=exogenous, trend='c', order=(1,1,0), seasonal_order=(1,1,1,52))
            #mod = sm.tsa.statespace.SARIMAX(endogenous, trend='n', order=(1,1,0), seasonal_order=(1,1,1,52))
            start_params = [0, 0, 1.]
            #results = mod.fit(start_params)
            results = mod.fit(start_params)#disp=False,maxiter=200)
            #Extract the coefficients
            coeff = list(results.params)[:5]
            """
            #print results.summary()#.tables[1]
            #predictions = results.predict(start = len(exogenous), end = len(exogenous), dynamic= False,typ='levels', exog=current_feat)#exog=exogenous[-1:])#current_feat)
            #predictions.round(0)
        demand_forecast.append(ml_res)    
        var_coef.append(coeff)
        mod_intercept.append(intercept)
        price_range.append(pr)
        current_features.append(current_feat)
        feat_indx.append(i)
        products.append(product)
    features_df = pd.concat(current_features, axis=0)
    demand_forecast
    var_coef
    mod_intercept
    features_df = features_df.reset_index(drop=True)
    features_df.index = feat_indx
    oDict = {"Products": products, "SKU":feat_indx,"Demand_Forecast" : demand_forecast, "Intercept": mod_intercept,"Variables_Coeffs" : var_coef, "Price_Range" : price_range}
    result_df = pd.DataFrame(oDict,columns=["Products","SKU","Demand_Forecast","Intercept","Variables_Coeffs","Price_Range"])
    result_df["Demand_Forecast"] = np.where(result_df["Demand_Forecast"] == 0,1,result_df["Demand_Forecast"])
    #Reset the index and replace it with the SKUs
    result_df = result_df.reset_index(drop=True)
    result_df = result_df.set_index('SKU')
    return result_df,features_df




import pandas as pd
from pandas import ExcelWriter
#import itertools
import time
#import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


dd_forecast = demand_fxn('md_30.xlsx',"KPI_locationsku.xlsx","Price Calculator.xlsx")
dd_forecast[0]#.iloc[:,:-2]

#book = load_workbook('demand_forecast.xlsx')
writer = pd.ExcelWriter('demand_forecast.xlsx', engine='openpyxl') 
#writer.book = book
#writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

dd_forecast[0].iloc[:,:-2].to_excel(writer,sheet_name='Next_Month_Demand',index=True)
writer.save()
dd_result = writer.save()#_kpi.save()
dd_result



dd_fxn = demand_fxn('md_30.xlsx',"KPI_locationsku.xlsx","Price Calculator.xlsx")
bpr = pd.read_excel("Price Calculator.xlsx", sheet_name='Sheet1')
#Reset the index and replace it with the SKUs
bpr = bpr.reset_index(drop=True)
bpr = bpr.set_index('ID')
curr_feat = dd_fxn[1]
coef = dd_fxn[0].iloc[:,-2:-1]
pr = dd_fxn[0].iloc[:,-1:]
pr
#dd_fxn[0]["Price_Range"]
intercept = dd_fxn[0].iloc[:,-3:-2]
ind_sk = sorted(list(dd_fxn[0].index.unique()))
optimal_prices = []
opt_sku = []
for i in ind_sk:
    price_range = pr.loc[i,"Price_Range"]
    if i in list(bpr.index):
        def profit_fxn(p_range):
            #ind_sk = sorted(list(dd_fxn.index.unique()))
            #for i in ind_sk:
            #p_range = pr.loc[i,"Price_Range"]
            op_prf = ((intercept.loc[i,"Intercept"]+(coef.loc[i,"Variables_Coeffs"][0]*p_range)
            +(coef.loc[i,"Variables_Coeffs"][1]*curr_feat.loc[i,"loc_per_SKU_shift"])+
            (coef.loc[i,"Variables_Coeffs"][2]*curr_feat.loc[i,"Seasons_shift"])
            +(coef.loc[i,"Variables_Coeffs"][3]*curr_feat.loc[i,"Trend_shift"]))*p_range)
            -bpr.loc[i][-1]#.loc[i,"Buy Price"]
            return -op_prf
        #Conduct a global optimization with the given price ranges
        #opt_price = spo.brute(profit_fxn,(price_range[1]-10,price_range[1]+20,1),finish=spo.fmin)
        opt_price = spo.fminbound(profit_fxn, price_range[0], price_range[1])
        optimal_prices.append(opt_price.round())
        opt_sku.append(i)
oDict = {"Optimum Price":optimal_prices,"SKU":opt_sku}#, "SKU":sku,"Demand_Forecast" : demand_forecast, "Intercept": mod_intercept,"Variables_Coeffs" : var_coef, "Price_Range" : price_range}
res_df = pd.DataFrame(oDict,columns=["Optimum Price","SKU"] )#,"Demand_Forecast","Intercept","Variables_Coeffs","Price_Range"])
res_df = res_df.set_index('SKU')
price_df = pd.concat([dd_fxn[0],res_df], axis=1, sort=False)
#price_df["Optimum Price"] = optimal_prices


price_df.to_excel(writer,sheet_name='Optimum_Prices',index=True)
dd_forecast[0].iloc[:,:-2].to_excel(writer,sheet_name='Next_Month_Demand',index=True)
writer.save()
dd_result = writer.save()#_kpi.save()
dd_result


bpr.loc[i][-1]
bpr.iloc[:,-1] = pd.to_numeric(bpr.iloc[:,-1],errors='coerce')

#spo.fminbound(f, 0, 10)

#Price Calculator.xlsx

bpr.loc[386,"Buy Price"]


  def opti_price(sr,mr,lr):
        """This function optimizes the mix of MA windows periods.
        sr is the range of short window values
        mr is the range of medium window values
        lr is the range of the long window values"""
        #Conduct a global optimization with the given ranges
        opt1 = spo.brute(ma_function,(sr,mr,lr),finish=None)
        #Conver the result to integers
        win_int = opt1.astype(int)
        #Conduct a local optimization with the given ranges
        #Remember to convert the windows from opt1 to integers
        #opt2 = spo.fmin(ma_function,win_int,xtol=1,ftol=1,maxiter=15,maxfun=20)
        return tuple(win_int)

print(results.mle_retvals)
    
    
    
    
    
    
    #Test the model using in-sample forecasts. 
    in_sample_forecasts = results.predict(start=1, end=len(exogenous)+len(exogenous)-1,typ='levels', exog=exogenous,dynamic=False)#.fillna(25.4)
    in_sample_forecasts = in_sample_forecasts.round(0)[:-16]
     
    
    #Test the model using in-sample forecasts. 
    oDict = {"Actual_Demand" : endogenous[1:], "Forecasted_Demand" : ml_res.round(0)}
    back_test = pd.DataFrame(oDict,columns=["Actual_Demand","Forecasted_Demand"])
    print back_test.head(10)
    plt.figure()
    plt.plot(back_test["Actual_Demand"],label="Actual_Demand", color='blue')
    plt.plot(back_test["Forecasted_Demand"], label="Forecasted_Demand",color='red',linewidth=2)
    plt.xlabel("dates")
    plt.ylabel("target")
    plt.title("Demand Forecast Result Test"+" "+"("+products[2]+")")
    plt.legend()
    plt.show()
        
    
    X = range(1,len(endogenous))#months
    y = endogenous[1:]
    #Predict
    y_2 = ml_res.round(0)
    
    # Plot the results
    plt.figure()
    plt.scatter(X, y, c="k", label="Actual_Demand")
    #plt.plot(X, y_1, c="g", label="n_estimators=1", linewidth=2)
    plt.plot(X, y_2, c="r", label="Forecasted_Demand", linewidth=2)
    plt.xlabel("data")
    plt.ylabel("target")
    plt.title("Boosted Decision Tree Regression")
    plt.legend()
    plt.show()
    
    
    from sklearn.linear_model import LinearRegression
    #X = np.array([[1, 1], [1, 2], [2, 2], [2, 3]])
    # y = 1 * x_0 + 2 * x_1 + 3
    #y = np.dot(X, np.array([1, 2])) + 3
    reg = LinearRegression().fit(exogenous, endogenous)
    reg.score(exogenous, endogenous)
    reg.coef_
    int(reg.intercept_)
    print reg.predict(current_feat)
   







#abr.score(exogenous, endogenous)#Returns the coefficient of determination R^2 of the prediction.
#0.36751657232999979



full_scores = cross_val_score(abr, exogenous, endogenous,cv=3)#estimate the accuracy of the model
full_scores.mean()
#Out[160]: array([ 0.02784944,  0.3762485 , -0.20821144])

#full_scor = cross_val_score(abr, exogenous, endogenous,cv=3, scoring = "neg_mean_squared_error")#estimate the accuracy of the model
#full_scor

X = range(1,len(map(str,list(dd_df.index.values))[1:])+1)#months
y = endogenous
#Predict
y_2 = abr.predict(exogenous)

# Plot the results
plt.figure()
plt.scatter(X, y, c="k", label="training samples")
#plt.plot(X, y_1, c="g", label="n_estimators=1", linewidth=2)
plt.plot(X, y_2, c="r", label="n_estimators=300", linewidth=2)
plt.xlabel("data")
plt.ylabel("target")
plt.title("Boosted Decision Tree Regression")
plt.legend()
plt.show()




#I am getting this error. Thanks.

#ValueError: Out-of-sample forecasting in a model with a regression component requires additional exogenous values via the exog argument.





def allocation_formula(prod_id):
    """ '''prod_id= the product ID; should be entered with a quote''' """
    #Read the excel file of the raw data and load it into a pandas dataframe
    data_sold = pd.read_excel('allocation_formula2.xlsx', sheetname='Sold')
    data_dispensed = pd.read_excel('allocation_formula2.xlsx', sheetname='Dispensed')
    sold = pd.DataFrame(data_sold)._get_numeric_data().fillna(0)
    dispensed = pd.DataFrame(data_dispensed)._get_numeric_data().fillna(0)
    sold.set_index('Product ID', inplace=True)
    dispensed.set_index('Product ID', inplace=True)
    del sold.index.name
    del dispensed.index.name
    sold = sold.T
    dispensed = dispensed.T
    #save dataframes as csv in the working directory
    sold.to_csv('drugs_sold_csv.csv')
    dispensed.to_csv('drugs_dispensed_csv.csv')
    sold_csv = pd.read_csv("drugs_sold_csv.csv", index_col='Unnamed: 0')
    dispensed_csv = pd.read_csv("drugs_dispensed_csv.csv", index_col='Unnamed: 0')
    sld = pd.to_datetime(sold_csv.index).date
    disp = pd.to_datetime(dispensed_csv.index).date
    sold_csv.index = sld
    dispensed_csv.index = disp
    #Set the index names to 'weeks'
    sold_csv.index.name = "Weeks"
    dispensed_csv.index.name = "Weeks"
    sold_df = pd.DataFrame(sold_csv)
    dispened_df = pd.DataFrame(dispensed_csv)
    total_sales = sold_df + dispened_df
    product = total_sales[prod_id]
    if product.mean() ==0:
        return "There are no forecasts available for this drug; no sales in the last six months. Try another drug"
    else:
        model = ARIMA(product, order=(2,1,1),freq='7D',missing = 'drop')
    model_fit = model.fit(disp=-1)
    model_fit = model_fit
    forecast_d = model_fit.predict(start=len(product), end=len(product), typ='levels',dynamic=False) 
    next_wk_forecast = forecast_d.round(0)
    return next_wk_forecast

allocation_formula('11')

   

"""
def allocation_formula3(prod_id):
    """ '''prod_id= the product ID; should be entered with a quote''' """
    #Read the excel file of the raw data and load it into a pandas dataframe and clean it
    data_sold = pd.read_excel('allocation_formula2.xlsx', sheetname='Sold')
    data_dispensed = pd.read_excel('allocation_formula2.xlsx', sheetname='Dispensed')
    #load the data to dataframe and get only numeric data
    sold = pd.DataFrame(data_sold)._get_numeric_data().fillna(0)
    dispensed = pd.DataFrame(data_dispensed)._get_numeric_data().fillna(0)
    #set Product ID as the index
    sold.set_index('Product ID', inplace=True)
    dispensed.set_index('Product ID', inplace=True)
    #delete the index name to make things simpler, also since it will be replaced later on
    del sold.index.name
    del dispensed.index.name
    #Transpose the dataframe in order for the Prodct IDs to be the column labels
    sold = sold.T
    dispensed = dispensed.T
    #save dataframes as csv in the working directory and re-read it as csv to make it more acceptable to
    #ARIMA
    sold.to_csv('drugs_sold_csv.csv')
    dispensed.to_csv('drugs_dispensed_csv.csv')
    sold_csv = pd.read_csv("drugs_sold_csv.csv", index_col='Unnamed: 0')
    dispensed_csv = pd.read_csv("drugs_dispensed_csv.csv", index_col='Unnamed: 0')
    #convert the datatime index to date
    sld = pd.to_datetime(sold_csv.index).date
    disp = pd.to_datetime(dispensed_csv.index).date
    sold_csv.index = sld
    dispensed_csv.index = disp
    #Label the index as 'weeks'
    sold_csv.index.name = "Weeks"
    dispensed_csv.index.name = "Weeks"
    #load the csv files to dataframe
    sold_df = pd.DataFrame(sold_csv)
    dispened_df = pd.DataFrame(dispensed_csv)
    #Add drugs sold and drugs dispensed to get the total sales
    total_sales = sold_df + dispened_df
    #Isolate the total sales for each drug and run it in ARIMA
    product = total_sales[prod_id]
    model = ARIMA(product, order=(2,1,1),freq='7D',missing = 'drop')
    model_fit = model.fit(disp=-1)
    #forecast next week’s delivery amount
    forecast_d = model_fit.predict(start=len(product), end=len(product)+2, typ='levels',dynamic=False)
    next_wk_forecast = forecast_d.round(0)    
    #return the forecast
    return next_wk_forecast


allocation_formula3('99')
"""

Two_mths_forecast = model_fit.predict(start=len(indx_data), end=len(indx_data)+2, exog=indx_data, typ='levels',dynamic=False)
print "Two Month Forecast"
print Two_mths_forecast

def mean_sq_error(ID):   
    """prod_id= the product ID; should be entered with a quote"""
    #Read the excel file of the raw data and load it into a pandas dataframe and clean it
    data_sold = pd.read_excel('allocation_formula2.xlsx', sheetname='Sold')
    data_dispensed = pd.read_excel('allocation_formula2.xlsx', sheetname='Dispensed')
    #load the data to dataframe and get only numeric data
    sold = pd.DataFrame(data_sold)._get_numeric_data().fillna(0)
    dispensed = pd.DataFrame(data_dispensed)._get_numeric_data().fillna(0)
    #set Product ID as the index
    sold.set_index('Product ID', inplace=True)
    dispensed.set_index('Product ID', inplace=True)
    #delete the index name to make things simpler, also since it will be replaced later on
    del sold.index.name
    del dispensed.index.name
    #Transpose the dataframe in order for the Prodct IDs to be the column labels
    sold = sold.T
    dispensed = dispensed.T
    #save dataframes as csv in the working directory and re-read it as csv to make it more acceptable to
    #ARIMA
    sold.to_csv('drugs_sold_csv.csv')
    dispensed.to_csv('drugs_dispensed_csv.csv')
    sold_csv = pd.read_csv("drugs_sold_csv.csv", index_col='Unnamed: 0')
    dispensed_csv = pd.read_csv("drugs_dispensed_csv.csv", index_col='Unnamed: 0')
    #convert the datatime index to date
    sld = pd.to_datetime(sold_csv.index).date
    disp = pd.to_datetime(dispensed_csv.index).date
    sold_csv.index = sld
    dispensed_csv.index = disp
    #Label the index as 'weeks'
    sold_csv.index.name = "Weeks"
    dispensed_csv.index.name = "Weeks"
    #load the csv files to dataframe
    sold_df = pd.DataFrame(sold_csv)
    dispened_df = pd.DataFrame(dispensed_csv)
    #Add drugs sold and drugs dispensed to get the total sales
    total_sales = sold_df + dispened_df
    #Isolate the total sales for each drug 
    prod = total_sales[ID]
    #create a size variable
    size = int(len(prod) * 0.66)
    #divide the data into training and testing components and run the ARIMA with the training data
    train, test = prod[0:size], prod[size:len(prod)]
    model = ARIMA(train, order=(2,1,1),freq='7D',missing = 'drop')
    model_fit = model.fit(disp=-1)
    #forecast the product amount for the period covering the testing data
    forecast_w = model_fit.predict(start=len(train), end=len(prod)-1, exog=train,
    typ='levels',dynamic=False)
    #run the mean_squared_error test and return its value 
    MSE = mean_squared_error(test, forecast_w)
    return MSE


mean_sq_error('2')



